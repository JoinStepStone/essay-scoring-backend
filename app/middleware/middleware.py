import jwt
from flask import request, jsonify, g
from functools import wraps
from app import user_database, app
import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd
from openpyxl.styles.colors import Color
from io import BytesIO
import math
import base64
import zipfile
from xml.etree import ElementTree as ET
from lxml import etree as ET
import re
# Secret key for JWT encoding/decoding
SECRET_KEY = "your_secret_key_here"
ALLOWED_EXTENSIONS = {'xlsm'}

def is_direct_cell_reference(formula):
    if isinstance(formula, str):
        # Remove the leading '='
        formula = formula.lstrip('=').strip()
        
        # Regular expression to match a cell reference optionally with sheet name
        pattern = r'[A-Za-z][0-9]'
        
        return re.match(pattern, formula) 

    return True

def fill_values_get_score(source_wbb, target_file):
    target_wb = openpyxl.load_workbook(target_file, keep_vba=True, data_only= True)
    source_wb = openpyxl.load_workbook(source_wbb, keep_vba=True, data_only= True)
    target_wb_with_formula = openpyxl.load_workbook(target_file, keep_vba=True)
    replace_name_dict = {"Toggle Valuation_Solution": "Valuation Model", "Toggle Model_Solutions": "Financial Model"}
    score = []
    total_score = []
    for sheet_name in ['Grading Key', 'Grading Key Sensitivity Table']:
        grading_key_sheet = source_wb[sheet_name]
        grading_key_sheet_with_formula = target_wb_with_formula[sheet_name]

        keep_cells = ["C"]
        not_found_value = True
        sensitive_value = 0
        for row_idx, row in enumerate(grading_key_sheet.iter_rows(), start=1):
            for cell in row:
                if cell.column_letter not in keep_cells:
                    continue
                elif cell.value is None:
                    continue
                elif "!" not in cell.value:
                    continue
                elif "Valuation_Solution" in cell.value:
                    target_sheet_name = replace_name_dict["Toggle Valuation_Solution"]
                else:
                    target_sheet_name = replace_name_dict["Toggle Model_Solutions"]

                if sheet_name == 'Grading Key Sensitivity Table' and not_found_value:
                    not_found_value = False
                    sensitive_value = grading_key_sheet['F'+str(row_idx - 1)].value
                    total_score.append(grading_key_sheet['F'+str(row_idx - 1)].value)

                if sheet_name != 'Grading Key Sensitivity Table':
                    total_score.append(grading_key_sheet['F'+str(row_idx)].value)

                cell_number = cell.value.split('!')[1]

                target_sheet = target_wb[target_sheet_name]  # Financial model or valuation model
                target_cell_value = target_sheet[cell_number].value # Pick the value from the cell
                grading_key_sheet['E'+str(row_idx)] = target_cell_value # Assign it to gradingkey cell E

                if target_cell_value is None:
                    target_cell_value = 0
                
                if'isformula' in grading_key_sheet_with_formula['G'+str(row_idx)].value.lower():
                    if is_direct_cell_reference(target_wb_with_formula[target_sheet_name][cell_number].value):
                        if sheet_name == 'Grading Key Sensitivity Table':
                            sensitive_value = 0
                        grading_key_sheet['G'+str(row_idx)] = 0
                    else:
                        if sheet_name != 'Grading Key Sensitivity Table':
                            score.append(grading_key_sheet['F'+str(row_idx)].value)
                        grading_key_sheet['G'+str(row_idx)] = grading_key_sheet['F'+str(row_idx)].value
                    continue
                if round(grading_key_sheet['D'+str(row_idx)].value,3) != round(target_cell_value,3):
                    if sheet_name == 'Grading Key Sensitivity Table':
                        sensitive_value = 0
                    # print(f"Row: {row_idx} Value in {cell_number} of sheet '{target_sheet_name}': {round(grading_key_sheet['D'+str(row_idx)].value,3)} {round(target_cell_value,3)} ")
                    grading_key_sheet['G'+str(row_idx)] = 0
                else:
                    if sheet_name != 'Grading Key Sensitivity Table':
                        score.append(grading_key_sheet['F'+str(row_idx)].value)
                    grading_key_sheet['G'+str(row_idx)] = grading_key_sheet['F'+str(row_idx)].value
                    
    output = BytesIO()
    source_wb.save(output)
    output.seek(0) 

    # Base64 encode the binary data for sending in JSON
    encoded_excel = base64.b64encode(output.read()).decode('utf-8')
    return {"score": round(((sum(score)+sensitive_value)/sum(total_score))*100,2), 'file': encoded_excel}, True, ""

def create_new_sheet_xml_with_data(data):
    # Define the namespaces
    nsmap = {
        None: 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
        'mc': 'http://schemas.openxmlformats.org/markup-compatibility/2006',
        'x14ac': 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac',
    }

    # Create the root element with namespace declarations
    worksheet = ET.Element('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}worksheet', nsmap=nsmap)
    worksheet.set('{http://schemas.openxmlformats.org/markup-compatibility/2006}Ignorable', 'x14ac')

    # Add required elements
    dimension = ET.SubElement(
        worksheet,
        '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}dimension',
        {'ref': f'A1:E{len(data) + 1}'}
    )
    sheet_views = ET.SubElement(
        worksheet,
        '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheetViews'
    )
    sheet_view = ET.SubElement(
        sheet_views,
        '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheetView',
        {'workbookViewId': '0'}
    )
    sheet_format_pr = ET.SubElement(
        worksheet,
        '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheetFormatPr',
        {
            'defaultRowHeight': '15',
            '{http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac}dyDescent': '0.25'
        }
    )

    sheet_data = ET.SubElement(
        worksheet,
        '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheetData'
    )

    def handle_nan(value):
        """Replace NaN with empty string or any other desired placeholder."""
        return "" if isinstance(value, float) and math.isnan(value) else value

    # First, add the header row (row 1)
    headers = ["Cell", "Solutions", "Student", "Credit Values", "Cell Score"]
    header_row = ET.SubElement(
        sheet_data,
        '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row',
        {'r': '1'}
    )
    for col_idx, header in enumerate(headers):
        col_letter = chr(ord('A') + col_idx)  # Column letters: A, B, C, D, E...
        cell_ref = f"{col_letter}1"
        cell_elem = ET.SubElement(
            header_row,
            '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c',
            {'r': cell_ref, 't': 'str'}
        )
        value_elem = ET.SubElement(
            cell_elem,
            '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v'
        )
        value_elem.text = header  # Set header text

    # Now, iterate through the data starting at row 2
    for row_idx, row_data in enumerate(data, 2):  # Row index starts from 2 (after the header)
        row_elem = ET.SubElement(
            sheet_data,
            '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row',
            {'r': str(row_idx)}
        )

        # Columns A to E
        columns = ['Cell', 'Solutions', 'Student', 'Credit Values', 'Cell Score']
        for col_idx, col_name in enumerate(columns):
            col_letter = chr(ord('A') + col_idx)
            cell_ref = f"{col_letter}{row_idx}"
            cell_type = 'n' if col_name != 'Cell' else 'str'
            cell_elem = ET.SubElement(
                row_elem,
                '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c',
                {'r': cell_ref, 't': cell_type}
            )
            value_elem = ET.SubElement(
                cell_elem,
                '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v'
            )
            value = handle_nan(row_data.get(col_name, ""))
            if cell_type == 'n' and value == "":
                value = "0"  # Default numerical value if NaN or empty
            value_elem.text = str(value)

    # Convert the XML tree to a string with proper encoding and XML declaration
    return ET.tostring(worksheet, encoding='utf-8', xml_declaration=True, pretty_print=False)

# Function to add a new sheet to the workbook
def add_new_sheet_to_workbook(workbook_xml_data, new_sheet_name, new_sheet_id, new_r_id):
    # Define namespaces
    nsmap = {
        'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    }

    # Parse XML with lxml
    root = ET.fromstring(workbook_xml_data)

    # Find the sheets element
    sheets_elem = root.find('.//main:sheets', namespaces=nsmap)
    if sheets_elem is None:
        # If sheets element is not found, create it
        sheets_elem = ET.SubElement(
            root,
            '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets'
        )

    # Create a new sheet element
    new_sheet = ET.Element(
        '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet',
        attrib={
            'name': new_sheet_name,
            'sheetId': str(new_sheet_id),
            '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id': new_r_id
        }
    )

    # Append the new sheet element
    sheets_elem.append(new_sheet)

    # Convert back to string
    return ET.tostring(root, encoding='utf-8', xml_declaration=True)

# Modify relationships file to add new sheet relationship
def add_new_sheet_relationship(rels_xml_data, new_r_id, new_sheet_number):
    # Parse XML with lxml
    root = ET.fromstring(rels_xml_data)

    # Create a new relationship element
    new_rel = ET.Element('Relationship', attrib={
        'Id': new_r_id,
        'Type': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet',
        'Target': f'worksheets/sheet{new_sheet_number}.xml'
    })

    # Append the new relationship
    root.append(new_rel)

    # Convert back to string
    return ET.tostring(root, encoding='utf-8', xml_declaration=True)

# Function to get the next sheet number
def get_next_sheet_number(workbook_xml_data):
    # Define namespaces
    nsmap = {
        'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    }

    # Parse XML with lxml
    root = ET.fromstring(workbook_xml_data)

    sheets = root.findall('.//main:sheet', namespaces=nsmap)
    if sheets:
        last_sheet_id = max(int(sheet.get('sheetId')) for sheet in sheets)
        return last_sheet_id + 1
    else:
        return 1  # If no sheets exist, start with sheet 1

# Function to add data sheets to the workbook
def add_data_sheets(df1, df2, current_workbook):
    output = BytesIO()

    with zipfile.ZipFile(current_workbook, 'r') as archive:
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as new_archive:
            # Read the workbook.xml to get the next sheet IDs
            workbook_xml_data = archive.read('xl/workbook.xml')
            new_sheet_id_1 = get_next_sheet_number(workbook_xml_data)
            new_sheet_id_2 = new_sheet_id_1 + 1
            new_r_id_1 = f'rId{new_sheet_id_1}'
            new_r_id_2 = f'rId{new_sheet_id_2}'

            # Modify workbook.xml
            workbook_xml_modified = add_new_sheet_to_workbook(
                workbook_xml_data, 'Grading Key Computed', new_sheet_id_1, new_r_id_1)
            workbook_xml_modified = add_new_sheet_to_workbook(
                workbook_xml_modified, 'Grading Key Sensitivity Table Computed', new_sheet_id_2, new_r_id_2)

            # Modify workbook.xml.rels
            rels_xml_data = archive.read('xl/_rels/workbook.xml.rels')
            rels_xml_modified = add_new_sheet_relationship(
                rels_xml_data, new_r_id_1, new_sheet_id_1)
            rels_xml_modified = add_new_sheet_relationship(
                rels_xml_modified, new_r_id_2, new_sheet_id_2)

            # Write all files, replacing modified ones
            for file_info in archive.infolist():
                file_data = archive.read(file_info.filename)

                if file_info.filename == 'xl/workbook.xml':
                    # Write modified workbook.xml
                    new_archive.writestr(file_info.filename, workbook_xml_modified)
                elif file_info.filename == 'xl/_rels/workbook.xml.rels':
                    # Write modified relationships
                    new_archive.writestr(file_info.filename, rels_xml_modified)
                else:
                    # Write unmodified file
                    new_archive.writestr(file_info, file_data)

            # Add the first new sheet file with data from df1
            new_sheet_xml_1 = create_new_sheet_xml_with_data(df1)
            new_archive.writestr(f'xl/worksheets/sheet{new_sheet_id_1}.xml', new_sheet_xml_1)

            # Add the second new sheet file with data from df2
            new_sheet_xml_2 = create_new_sheet_xml_with_data(df2)
            new_archive.writestr(f'xl/worksheets/sheet{new_sheet_id_2}.xml', new_sheet_xml_2)

    # Reset the buffer's position to the beginning for reading
    output.seek(0)

    return output

def copy_sheet(source_file, target_file,keep_values = False):
    source_wb = openpyxl.load_workbook(source_file, keep_vba=True, data_only=True)
    sheet_names = source_wb.sheetnames
    sheet_num = 1
    data = []
    numbers = []
    for sheet_name in sheet_names:
        if sheet_name in ['Grading Key Sensitivity Table', 'Grading Key']:
            data.append(get_df(source_file,sheet_name) )
            numbers.append(sheet_num)
        sheet_num += 1
    
    # data[0].insert(0, {"Cell":"Cell", "Solutions": "Solutions", "Student": "Student", "Credit Values": "Credit Values", "Cell Score": "Cell Score"})
    # data[1].insert(0, {"Cell":"Cell", "Solutions": "Solutions", "Student": "Student", "Credit Values": "Credit Values", "Cell Score": "Cell Score"})
    # print("\n", data[0][0], data[0][1], "\n")
    return add_data_sheets(data[0], data[1], target_file)
    
    

def paste_extracted_df(data, file, sheet_name):
    wb = openpyxl.load_workbook(file)

    # Create a new sheet or select an existing one
    new_sheet_name = sheet_name
    if new_sheet_name in wb.sheetnames:
        sheet = wb[new_sheet_name]
    else:
        sheet = wb.create_sheet(title=new_sheet_name)

    # Write the headers based on the keys of the first dictionary
    headers = list(data[0].keys())
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Writing the list of dictionaries into the new sheet
    for row_num, entry in enumerate(data, 2):  # Start from the second row for data
        for col_num, key in enumerate(entry.keys(), 1):
            value = entry[key]
            
            # Handle NaN (or None) values, replace with an empty string or any placeholder
            if isinstance(value, float) and math.isnan(value):
                value = ''
            
            # Write the value into the corresponding cell
            print("\n", "Column: ", col_num, " Value: ", value, "\n")
            sheet.cell(row=row_num, column=col_num, value=value)

    # Save the updated workbook
    wb.save("path_to_your_file.xlsm")

def trim_dict(d, max_keys=5):
    # Convert the dictionary to a list of tuples, and slice to get the first 5 items
    return dict(list(d.items())[:max_keys])

def get_df(source_file, sheet_name):

    df = pd.read_excel(source_file, sheet_name=sheet_name, engine='openpyxl')
    df_cleaned = df.dropna(how='all', axis=1)
    
    if df_cleaned.iloc[0].notna().sum() > 1:  
        df_cleaned.columns = df_cleaned.iloc[0]
        df_cleaned = df_cleaned.drop(0).reset_index(drop=True)
    
    df_cleaned = df_cleaned.dropna(how='all')  # Remove rows where all values are NaN
    df_cleaned = df_cleaned.reset_index(drop=True)
    data_dict = df_cleaned.to_dict(orient='records')  # Use 'records' to get a list of dictionaries

    # Apply the trimming function to each dictionary in the list
    trimmed_data = [trim_dict(d) for d in data_dict]
    
    return trimmed_data

def compare_results(file, data_dict):
    workbook = openpyxl.load_workbook(file, keep_vba=True)

    for data_object in data_dict:
        # Access the target sheet
        sheet_name = data_object['target_file_name']

        # Check if the sheet exists in the workbook
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            return "", False, "Missing Target File"

        # Access the specific cell mentioned in the object
        cell_number = data_object['cell_number']
        cell_value = sheet[cell_number].value
        print("Cell Value: ", cell_value, cell_number)

    return 88, True, ""

def get_cell_value(file, sheet_name, cell):
    
    df = pd.read_excel(file, sheet_name=sheet_name, header=None)
    
    column_letter = cell[0]  
    row_number = int(cell[1:]) - 1
    
    # Convert column letter to a zero-based index
    column_index = ord(column_letter.upper()) - ord('A')
    
    # Retrieve the value from the DataFrame
    cell_value = df.iloc[row_number, column_index]
    
    return cell_value

def make_dict_grading_key(value_list):
    replace_name_dict = {"Toggle Valuation_Solution": "Valuation Model", "Toggle Model_Solutions": "Financial Model"}
    if len(value_list[0]) == len(value_list[1]) == len(value_list[2]):
        dictionary_grading_list = []
        value_list[0].pop(0)
        value_list[1].pop(0)
        value_list[3].pop(0)

        for cell_value1, cell_value2, cell_value3 in zip(value_list[0], value_list[1], value_list[5]):
            try:
                if cell_value1 == "Not a formula":
                    continue
                if isinstance(cell_value1, str):
                    cell_value_cleaned = cell_value1.strip("=")
                    file_name, cell_reference = cell_value_cleaned.split('!')
                    if(cell_reference == "L40"):
                        print(cell_value3)
                    file_name_clean = file_name.strip("'")
                    parts = file_name_clean.split('_')
                    parts[1] = "Student"
                    modified_string = '_'.join(parts)
                    dictionary_grading_list.append({
                        "file_name" : file_name_clean,
                        "cell_number" : cell_reference,
                        "cell_value": cell_value2,
                        "grade": cell_value3,
                        "target_file_name": replace_name_dict[file_name_clean]
                    })
            except Exception as e:
                print("Error: ", str(e))
                return "", False, "Error in processing"

        return dictionary_grading_list, True, ""

    return "", False, "Grading Key Workbook has different length of cell, values and grades"

def remove_unnessary_data(grading_key_df):
    # Drop rows where all elements are NaN and drop columns that are entirely NaN
    grading_key_df.dropna(how='all', inplace=True)
    grading_key_df.dropna(axis=1, how='all', inplace=True)
    
    # Convert to dictionary format
    grading_key_dict = grading_key_df.to_dict(orient='list')

    return grading_key_dict

def parsed_xlsx_get_score(file):

    xlsx = pd.ExcelFile(file, engine='openpyxl')
    
    # Check if 'Grading Key' sheet exists
    if 'Grading Key' in xlsx.sheet_names:
        # Parse the 'Grading Key' sheet
        grading_key_df = pd.read_excel(file, sheet_name='Grading Key')
        cleaned_original_data = remove_unnessary_data(grading_key_df)
        dictionary_grading_key, success, error = make_dict_grading_key(list(cleaned_original_data.values()))

        return dictionary_grading_key, success, error
    else:
        return "", False, "Grading Key sheet not found in the Excel file."


def upload_file(file):
    fileDummy = file.filename.rsplit('.', 1)[0]
    extension = file.filename.rsplit('.', 1)[1]
    filename = fileDummy + str(datetime.datetime.utcnow()) + "." + extension
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    # Process the string_field and file here
    return filepath

def generate_access_token(info, expires_in="12h"):
    # Convert expires_in to seconds (12 hours = 43200 seconds)
    if expires_in == "12h":
        expiration = datetime.datetime.utcnow() + datetime.timedelta(hours=12)
    else:
        raise ValueError("Unsupported expiration format")

    # Generate the JWT
    token = jwt.encode(
        {"info": info, "exp": expiration},
        SECRET_KEY,
        algorithm="HS256"
    )

    return token

def validate_token_duration(request):
    try:
        # Extract the token from the Authorization header
        auth_header = request.headers.get('Authorization', None)
        if auth_header is None:
            return {"data": "", "code": 401, "message": "Header does not have token"}

        token = auth_header.split(" ")[1]
        
        # Decode the JWT token
        try:
            jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        except jwt.ExpiredSignatureError:
            return "", False, "Token has been expired"
        except jwt.InvalidTokenError:
            return "", False, "Token is invalid"

        return "", True, "Token is valid"
    except Exception as e:
        return "",False, str(e)

def validate_token_admin(f):
    @wraps(f)
    def decorated_function_admin(*args, **kwargs):
        try:
            # Extract the token from the Authorization header
            auth_header = request.headers.get('Authorization', None)
            if auth_header is None:
                return {"data": "", "code": 401, "message": "Header does not have token"}

            token = auth_header.split(" ")[1]
            
            # Decode the JWT token
            try:
                data = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
            except jwt.ExpiredSignatureError:
                return {"data": "", "code": 401, "message": "Token has been expired"}
            except jwt.InvalidTokenError:
                return {"data": "", "code": 401, "message": "Token is invalid"}

            # Find the user in the database
            user = user_database.find_one({"email": data['info']["email"], "role": "Admin"})
            if not user:
                return {"data": "", "code": 401, "message": "User is not found"}

            # Attach user info to the global `g` object
            g.current_user = {
                "email": user["email"],
                "role": user["role"],
                "_id": str(user["_id"]),
                "name": user.get("name")
            }

            return f(*args, **kwargs)
        except Exception as e:
            return {"data": "", "code": 401, "message": str(e)}

    return decorated_function_admin

def validate_token(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        try:
            # Extract the token from the Authorization header
            auth_header = request.headers.get('Authorization', None)
            if auth_header is None:
                return {"data": "", "code": 401, "message": "Header does not have token"}

            token = auth_header.split(" ")[1]

            # Decode the JWT token
            try:
                data = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
            except jwt.ExpiredSignatureError:
                return {"data": "", "code": 401, "message": "Token has been expired"}
            except jwt.InvalidTokenError:
                return {"data": "", "code": 401, "message": "Token is invalid"}

            # Find the user in the database
            user = user_database.find_one({"email": data['info']["email"], "role": "Student"})
            if not user:
                return {"data": "", "code": 401, "message": "User is not found"}

            # Attach user info to the global `g` object
            g.current_user = {
                "email": user["email"],
                "role": user["role"],
                "_id": str(user["_id"]),
                "name": user.get("name")
            }

            return f(*args, **kwargs)
        except Exception as e:
            return {"data": "", "code": 401, "message": str(e)}

    return decorated_function

# Function to get the current user from `g`
def get_current_user():
    return getattr(g, "current_user", None)

def remove_sheets(current_workbook):
    # workbook = openpyxl.load_workbook(current_workbook, keep_vba=True, data_only=True)
    sheets_to_keep = ['Instructions', 'Financial Model', 'Valuation Model']

    output = BytesIO()
    with zipfile.ZipFile(current_workbook, 'r') as archive:
        with zipfile.ZipFile(output, 'w', compression=zipfile.ZIP_DEFLATED) as new_archive:
            for file_info in archive.infolist():
                file_data = archive.read(file_info.filename)

                if file_info.filename == 'xl/workbook.xml':
                    workbook_xml = ET.fromstring(file_data)
                    namespaces = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    
                    sheets = workbook_xml.findall('.//ns:sheet', namespaces)
                    
                    # Modify the sheet visibility for sheets not in sheets_to_keep
                    for sheet in sheets:
                        sheet_name = sheet.get('name')
                        if sheet_name not in sheets_to_keep:
                            sheet.set('state', 'veryHidden')  # Set visibility to 'veryHidden'
                    
                    # Convert the modified XML back to a byte string with proper encoding and declaration
                    file_data = ET.tostring(workbook_xml, encoding='utf-8', xml_declaration=True, standalone=False)
                
                # Write the (possibly modified) file back into the new ZIP archive
                new_archive.writestr(file_info, file_data)

    # Reset the buffer's position to the beginning for reading
    output.seek(0)

    # Return the modified Excel file as a byte stream
    return output

def visible_sheets(current_workbook):
    # workbook = openpyxl.load_workbook(current_workbook, keep_vba=True, data_only=True)

    # for sheet in workbook.sheetnames:
    #     std = workbook[sheet]
    #     std.sheet_state = 'visible'

    output = BytesIO()
    with zipfile.ZipFile(current_workbook, 'r') as archive:
        with zipfile.ZipFile(output, 'w', compression=zipfile.ZIP_DEFLATED) as new_archive:
            for file_info in archive.infolist():
                file_data = archive.read(file_info.filename)

                if file_info.filename == 'xl/workbook.xml':
                    workbook_xml = ET.fromstring(file_data)
                    namespaces = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    
                    sheets = workbook_xml.findall('.//ns:sheet', namespaces)
                    
                    # Modify the sheet visibility for sheets not in sheets_to_keep
                    for sheet in sheets:
                        sheet.set('state', 'visible')  # Set visibility to 'veryHidden'
                    
                    # Convert the modified XML back to a byte string with proper encoding and declaration
                    file_data = ET.tostring(workbook_xml, encoding='utf-8', xml_declaration=True, standalone=False)
                
                # Write the (possibly modified) file back into the new ZIP archive
                new_archive.writestr(file_info, file_data)

    # Reset the buffer's position to the beginning for reading
    output.seek(0)

    # Return the modified Excel file as a byte stream
    return output
    
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

